import pandas as pd
import os
from openpyxl import load_workbook
# from Config import UPLOAD_FOLDER, DUPLICATES_FOLDER, OUTPUT_DIR, DATA, GLOBAL_CONFIG, PAYLOAD_FILE,CLASSES
from Config import UPLOAD_FOLDER, OUTPUT_DIR, DATA, GLOBAL_CONFIG
from azure_Asyncnew import doc_extract, get_full_text_from_pagewise,doc_extract_single
from pathlib import Path
import itertools
import difflib as dl
#import duplicates as dup
import math
import copy

def get_meta_data_req(df,doc_claasifier):
    req_doc_meta_data=df[df["Doc Classification"]==doc_claasifier]
    if len(req_doc_meta_data)>0:
        return req_doc_meta_data
    else:
        return ''

def extract_metadata_info(meta_data_df,data):

    if len(meta_data_df)>0:
        record=meta_data_df.to_dict('records')
        text = get_full_text_from_pagewise(data)
        for i in record:
            regex_ex=i['regex']
            func_name=i['Function Name']
            keyword_name=i['Keyword/column name'] 
            metada_name=i['Metadata to be captured']
            
            if i["Extra Tag"] == 'Yes':
                regex_ex,func_name,keyword_name =  overwrite_metadata_config(data = text,regex = regex_ex,func_name = func_name,keyword_name=keyword_name)
            
            lib_import=__import__("post_processing")
            func=getattr(lib_import,func_name)
            data=func(data,metada_name,keyword_name,regex_ex)
    else:
        data['metadata_info']={}
    return data
  
def metatadata_save(metadata_dict:dict,file_name:str,**kwargs):
    if "sheet_name" in kwargs:
        sheet_name=kwargs["sheet_name"]
    else:
        sheet_name=''
    class_name=kwargs["class_name"]
    df=pd.DataFrame([metadata_dict]).T.reset_index()
    df.columns=['Metadata_name',"Metadata_value"]
    
    file_path=os.path.join(os.getcwd(),file_name)
  
    file_metadata_df=pd.read_excel(file_path)
    file_metadata_df['Category']=class_name

    writer=pd.ExcelWriter(file_name, engine = 'xlsxwriter')
  
    if sheet_name:
    
        file_metadata_df.to_excel(writer,index=False,sheet_name="File info")
        df.to_excel(writer,index=False,sheet_name=sheet_name)
        
        
    else:
        
        file_metadata_df.to_excel(writer,index=False,sheet_name="File info")
        df.to_excel(file_name,index=False)
    # writer.save()
    writer.close()
    remove_sheet(file_path,sheet_name="Sheet1")
    print("file is saved")
'''
def metadata_ocr_similarity(enitity_dict:list):
    match_list=[]
    final_match_list=[]
    id=0
    
    for i in enitity_dict:
        temp_dict={}
        category=i['Category']
        text=i['Text']
        file=i['file']
        match_list=[]
        temp_list=[]
        final_name =file.replace('pdf','xlsx')
        all_doc_dict=get_cat_file(file_path=CLASSES,category=category)
       
        if all_doc_dict:
            for m in all_doc_dict:
                
                result = round(dl.SequenceMatcher(None, all_doc_dict[m],text).ratio()*100,2)
                text_data=[text,all_doc_dict[m]]
                temp_dict.update({"filename":file,"existing_file":m,"match_score":result,"category":category,"text_data":text_data.copy()})
                match_list.append(temp_dict.copy())
        else:
            temp_dict.update({"filename":file,"existing_file":"","match_score":0,"category":category,"text_data":""})
            match_list.append(temp_dict.copy())
        
        for k in match_list:
            if k['match_score']>=95:
                
                temp_list.append({
                "ID": id,
                "Files":[k['filename'],k['existing_file']],
                "Text": k['text_data'],
                "Category": k['category'],
                "Similarity": k['match_score'],
                "output_file": '.'.join(list(map(lambda x: x.replace('pdf', 'xlsx'), k['filename'].split('.'))))
            })
                id+=1
        if temp_list:
            final_match_list.extend(temp_list.copy())
            
            metadata_save_in_file(final_list=temp_list,file_name=final_name)
        elif not temp_list:
            metadata_save_in_file(final_list=temp_list,file_name=final_name)
            for k in match_list:
                if k['match_score'] < 95:
                    temp_list.append({
                    "ID": id,
                    "Files":[k['filename']] ,
                    "Text": '' ,
                    "Category": k['category'],
                    "Similarity": '' ,
                    "output_file": '.'.join(list(map(lambda x: x.replace('pdf', 'xlsx'), k['filename'].split('.'))))
                })
           
            final_match_list.extend(temp_list.copy())
    
    return final_match_list

def get_cat_file(file_path:str,category:str):
    file_path=os.path.join(os.getcwd(),file_path,category)
    doc_dict={}
    if os.path.isdir(file_path) and os.listdir(Path(file_path))  : 
        for filename in os.listdir(Path(file_path)):
            upload_dir_path=str(file_path)+"/"+filename
            ocr_extraction = doc_extract([upload_dir_path])
            text = get_full_text_from_pagewise(ocr_extraction[upload_dir_path])
            doc_dict[filename]=text
    

    return doc_dict
'''

def metadata_save_in_file(final_list,file_name):
    file_name_op=os.path.join(os.getcwd(),OUTPUT_DIR, file_name)
    book = load_workbook(file_name_op)
    writer=pd.ExcelWriter(file_name_op, engine = 'openpyxl')
    writer.book = book
    if final_list:
        
        final_df=pd.DataFrame(final_list)
        final_df=final_df.drop("ID",axis=1).reset_index()
        final_df=final_df.drop("output_file",axis=1).reset_index()
        final_df.rename(columns = {'index':'ID'}, inplace = True)
        writer=pd.ExcelWriter(file_name_op, engine = 'openpyxl')
        writer.book = book
        final_df.to_excel(writer,index=False,sheet_name="Ocr Similarity")
        writer.close()
        print("file is saved")
    else:
     
        final_df=pd.DataFrame()
        final_df.to_excel(writer,index=False,sheet_name="Ocr Similarity")
        writer.close()
        print("file is saved")

'''
def hash_status_check(filename):
    upload_dir_path=str(UPLOAD_FOLDER)+"/"+filename
    upload_file_path=os.path.join(os.getcwd(),upload_dir_path)
    upload_hash=dup.hashfile(upload_file_path)
    dir_list = os.listdir(Path(CLASSES))
    file_size = os.path.getsize(upload_file_path)
    file_size_kb = math.ceil(file_size/1024)
    duplicate_list=[]
    for dir in dir_list:
        dir_path=str(CLASSES)+"/"+dir
        dir_full_path=os.path.join(os.getcwd(),dir_path)
        file_list = os.listdir(Path(dir_full_path))
        for class_file in file_list:
            class_dir_path=dir_path+"/"+class_file
            class_file_path=os.path.join(os.getcwd(),class_dir_path)
            class_hash=dup.hashfile(class_file_path)
            if str(upload_hash)==str(class_hash):
                duplicate_dict={"filename":filename,"filesize (KB)":file_size_kb,"Category":str(dir),"duplicate base":"Hash value","existing_filename":class_file,"existing_file_path":class_file_path}
                duplicate_list.append(duplicate_dict)
    return duplicate_list
'''

def remove_sheet(file_path,sheet_name):
    wb = load_workbook(file_path)
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    wb.save(file_path)


def get_file_paths_list(dir_path_list):
    list_files_path=[]
    for dir_path in dir_path_list:
        for list_dir in os.listdir(dir_path):
            abs_path=os.path.join(dir_path,list_dir)
            if os.path.isdir(abs_path):
                for file in os.listdir(abs_path):
                    abs_file_path=os.path.join(abs_path,file)
                    list_files_path.append(abs_file_path)
            else:
                
                list_files_path.append(abs_path)
    return list_files_path


def filer_file(new_file_path_list:list,all_file_path:list,threshold:int):
    filter_dict={}
    systme_diag_list=[]
    id=0
    for new_file_path in new_file_path_list:
        temp_list=[]
        # new files which needs to be processed 

        file_size = os.path.getsize(new_file_path)
        file_size=math.ceil(file_size/1024)
        new_file_path=new_file_path.replace('\\','/')
        new_file_name = new_file_path.split('/')[-1]
        max_range = file_size + file_size * (threshold/100)
        min_range = file_size - file_size * (threshold/100)
        dup_found_flag=False
        id+=1
        group_name="DAG"+str(id)
        # iterating over all files present in the given directory
        for old_file in all_file_path:
            old_file_size = os.path.getsize(old_file)
            old_file_size=math.ceil(old_file_size/1024)
            old_file_upd=old_file.replace('\\','/')
            old_file_name = old_file_upd.split('/')[-1]

            temp_dict={}
            
      
            if (old_file_size == file_size  or new_file_name == old_file_name) and threshold ==0 and new_file_path !=old_file_upd :
                temp_list.append(old_file)
                temp_dict.update({"Group Name":group_name,"File Name":new_file_name,"File Size":file_size,"Existing File":old_file_name,"Existing File Size": old_file_size ,"Existing File Path":old_file,\
                                  "Diagnostic":"likely to be duplicated because of same size","File Size var(%)":threshold})
                dup_found_flag=True
            # if threshold define to check on the range of size of file
            elif  threshold >0 and ((old_file_size >= min_range and old_file_size <= max_range) or new_file_name == old_file_name) and new_file_path !=old_file_upd :
                
                temp_list.append(old_file)
                temp_dict.update({"Group Name":group_name,"File Name":new_file_name,"File Size":file_size,"Existing File":old_file_name,"Existing File Size": old_file_size ,"Existing File Path":old_file,\
                                "Diagnostic":"likely to be duplicated because of same size","File Size var(%)":threshold})
                dup_found_flag=True
            if temp_dict:
                systme_diag_list.append(temp_dict.copy())
        else:
            if not dup_found_flag:
                temp_dict.update({"Group Name":"DAG0","File Name":new_file_name,"File Size":file_size,"Existing File":"","Existing File Size": "" ,"Existing File Path":"",\
                                "Diagnostic":"no duplicate found","File Size var(%)":threshold})
                systme_diag_list.append(temp_dict.copy())
                     
        
        
        filter_dict.update({new_file_path:temp_list.copy()})

    return filter_dict,systme_diag_list


def get_category(ocr_extract_dict:dict,config_data:pd.DataFrame,final_output_dict:dict):
    
    for doc_extraction in ocr_extract_dict:
        text = get_full_text_from_pagewise(ocr_extract_dict[doc_extraction])
        clean_path=doc_extraction.replace('\\','/')
        file_name=clean_path.split('/')[-1]
        for index, row in config_data.iterrows():
            keywords = row['Keywords']
            category = row['Category']
            
            keywords_list = keywords.split(',')
            for words in keywords_list:
                if words.lower() in text.lower():
                    category = row['Category'] if row['Category'] else "None"
                    ocr_extract_dict[doc_extraction].update({"category":category})
                    final_output_dict[file_name].update({"category":category})
                    break
            else:
                continue
            break

    return final_output_dict,ocr_extract_dict
'''
def get_metadata_extraction_wrapper(metadata_config:pd.DataFrame ,ocr_extraction:dict,final_output:dict):
    ocr_extraction_copy=copy.deepcopy(ocr_extraction)
    print(ocr_extraction)
    for file,data_dict in ocr_extraction.items():
        category=data_dict['category']
        print(category)
        # file_path_upd=file.replace('\\','/')
        # file_name=file_path_upd.split('/')[-1]
        metadata_df=get_meta_data_req(df = metadata_config,doc_claasifier = category)
        ocr_extraction_copy[file_name]=extract_metadata_info(meta_data_df =metadata_df,data=data_dict)
        final_output[file_name].update({"metadata_info":ocr_extraction_copy[file_name]['metadata_info']})
    return final_output
'''
def ocr_similarity_check_batch(all_file_data,new_file_data,final_output:dict):
    for file,data in new_file_data.items():
        temp_file_name=[]
        temp_file_similarity=[]
        file_path_upd=file.replace('\\','/')
        file_name=file_path_upd.split('/')[-1]
        new_file_full_text=get_full_text_from_pagewise(data)

        for exist_file, old_file_data in all_file_data.items():
             old_file_full_text=get_full_text_from_pagewise(old_file_data)
             result = round(dl.SequenceMatcher(None, new_file_full_text,old_file_full_text).ratio()*100,2)
             exist_file_upd=exist_file.replace('\\','/')
             exist_file_name=exist_file_upd.split('/')[-1]
             if result > 95 and exist_file_upd != file_path_upd:
                 
                 temp_file_name.append(exist_file_name)
                 temp_file_similarity.append(result)

        if temp_file_name:
            final_output[file_name].update({"ocr_similar_files":temp_file_name,"Similarity":temp_file_similarity})
        else:
            final_output[file_name].update({"ocr_similar_files":'',"Similarity":''})

    return final_output

def saving_files(final_output_dict:dict):
    output_file_path = os.path.join(os.getcwd(),OUTPUT_DIR,"batch_output.xlsx")
    
    output_list=[i for _,i in final_output_dict.items()]
    output_df=pd.DataFrame(output_list)
    output_df.to_excel(output_file_path)

def confidence_score(ocr_extraction,final_output,conifg_dict):

    for file,data in ocr_extraction.items():
        file_path_upd=file.replace('\\','/')
        file_name=file_path_upd.split('/')[-1]
        num_of_page=len(data)
        if num_of_page <= int(conifg_dict['page']) and all( data != ''  for i,data in final_output[file_name]['metadata_info'].items() ):
            final_output[file_name]['Confidence Level']="high"
        # elif num_of_page > int(conifg_dict['page']) and all( data != ''  for i,data in final_output[file_name]['metadata_info'].items() ):
        #     final_output[file_name]['Confidence Level']="medium"
        else:
            final_output[file_name]['Confidence Level']="medium"
    return final_output

def dict_to_list(final_ouput_dict):
    final_list=[]
    for _,data in final_ouput_dict.items():
        final_list.append(data)
    return final_list

def overwrite_metadata_config(data,regex,func_name,keyword_name):

    extra_tag_df=pd.read_excel(GLOBAL_CONFIG,sheet_name="Extra Tag").fillna('')
    extra_tag_recrds=extra_tag_df.to_dict('records')
    new_key=keyword_name
    new_function_name=func_name
    new_regex=regex
    for i  in extra_tag_recrds:

        key_word= i['Keyword/column name'].strip()
  
        if key_word  in data:
            
            new_regex=i['regex']
            new_function_name = i['Function Name']  if i['Function Name'] else new_function_name
            new_key =  i['Keyword/column name']  if i['Keyword/column name'] else new_key
       
            break

    return new_regex,new_function_name,new_key

def drop_redundant_rows(redundant_dict,sys_diag_df):
    index_list=[]
    for group,file_name in redundant_dict.items():
        index =list(sys_diag_df.loc[(sys_diag_df['Group Name'] == group) & (sys_diag_df['Existing File'] == file_name)].index)
        index_list.extend(index)
    sys_diag_df.drop(index_list,inplace=True)
    sys_diag_df.to_dict('records')
    return sys_diag_df

def get_redundant_file(sys_diag:list):
    sys_diag_df= pd.DataFrame(sys_diag)
    unique_grp=sys_diag_df["Group Name"].unique()
    redundant_dict={}
    for i in unique_grp:
        
        a_dict=sys_diag_df.loc[sys_diag_df['Group Name']==i].to_dict('records')
        
        for m in a_dict:
            file_name=m['File Name']
            if m['Existing File']:
                
                existing_file_list= list(sys_diag_df.loc[sys_diag_df['File Name'] == m['Existing File'],'Existing File'])
            
                if file_name in existing_file_list:
                    group_name=sys_diag_df.loc[sys_diag_df['File Name'] == m['Existing File'],'Group Name'].values[0]
                    redundant_dict[group_name] = file_name
        
        drop_redundant_rows(redundant_dict,sys_diag_df)
    return sys_diag_df